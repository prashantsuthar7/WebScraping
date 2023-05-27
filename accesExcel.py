from openpyxl import load_workbook

wb = load_workbook("P:/Destop/ZIP/ZipFolder/Zip_Code_missing.xlsx")  # Work Book
ws = wb.get_sheet_by_name('New')  # Work Sheet
column = ws['B']  # Column
column_list = [column[x].value for x in range(len(column))]

print(column_list)