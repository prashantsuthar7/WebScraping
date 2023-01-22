from requests_html import HTMLSession
from openpyxl import load_workbook
import csv
import time

wb = load_workbook("C:/Users/prash/Desktop/ZIP/ZipFolder/ZZip_Code_missing.xlsx")  # Work Book
ws = wb.get_sheet_by_name('New')  # Work Sheet
column = ws['B']  # Column
column_list = [column[x].value for x in range(len(column))]
#print(column_list)

s = HTMLSession()

extract_zip = []

#adds = ['1401 SOUTH DON ROSER DRIVE SUITE F1,LAS CRUCES']

for add in column_list:
    #print(add)
    url = 'https://www.google.com/search?q='+add
    r = s.get(url)
    try:
        zips = r.html.find('span.desktop-title-subcontent', first=True).text.strip()
    except AttributeError as err:
        zips = 'None'
        print(err)
    dis_zips = {

        'logic': add,
        'zip': zips
    }
    extract_zip.append(dis_zips)
    time.sleep(3)

    with open('result.csv', 'w', encoding='utf8',newline='') as f:
        wr = csv.DictWriter(f, fieldnames=extract_zip[0].keys(),)
        wr.writeheader()
        wr.writerows(extract_zip)
