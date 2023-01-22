from openpyxl import load_workbook
import requests
import pandas as pd
import os
from bs4 import BeautifulSoup


wb = load_workbook("C:/Users/prash/Desktop/ZIP/ZipFolder//Redfin Data Pull v2022.12.29.xlsx")  # Work Book
ws = wb.get_sheet_by_name('redfin_2022-12-29-15-00-14 (1)')  # Work Sheet
column = ws['U']  # Column
links = [column[x].value for x in range(len(column))]

#urls = ['https://www.redfin.com/IL/Winnetka/570-Oak-St-60093/home/13787294','https://www.redfin.com/IL/Wilmette/2128-Chestnut-Ave-60091/home/13784776']

#for link in links:
print(links)

header = {
  "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36",
  "X-Amzn-Trace-Id": "Root=1-62cc4e82-3e17734461ed9f8237c684d1"}

page_details = []

for url in links:
  r = requests.get(url, headers=header)
  soup = BeautifulSoup(r.content, 'lxml')
  print(url)
  try:
      listed_by = soup.find('div', class_ = 'listing-agent-item').text.strip()
  except:
      listed_by = None
  try:
      bought_with = soup.find('div', class_ = 'buyer-agent-item').text.strip()
  except:
      bought_with = None

  page_extract_details = {
      'URL' : url,
      'Listed By' : listed_by,
      'Bought with' : bought_with
  }
  page_details.append(page_extract_details)
#print(page_details)
df = pd.DataFrame(page_details)
output_path = 'redfin.com.csv'
df.to_csv(output_path, mode='a', header=not
os.path.exists(output_path)
          )